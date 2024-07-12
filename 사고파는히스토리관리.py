from openpyxl import load_workbook


class Main():
    def __init__(self):
        sCode = '001440, 232140'
        filePath = 'historyXlsx/'

        #self.xlsxDelete(sCode, filePath+'buyHistory.xlsx')
        self.xlsxAdd(sCode, filePath+'buyHistory.xlsx')


    def xlsxDelete(self, sCode, fName):
        print('sCode : %s' % sCode)
        # 엑셀 파일 불러오기
        workbook = load_workbook(fName)

        # 첫 번째 시트 선택
        sheet = workbook.active

        # 'buyHistory' 컬럼 내에서 sCode가 있는 행 찾기
        for idx, row_data in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            print('비교 - %s : %s' % (row_data[1], sCode))
            if row_data[1] == sCode:  # 'buyHistory' 컬럼의 인덱스가 0이라고 가정
                # 해당 행 삭제
                print(row_data[1])
                sheet.delete_rows(idx)

        # 변경된 내용을 저장
        workbook.save(fName)

    def xlsxAdd(self, sCode, fName):
        print('sCode : %s' % sCode)

        # 엑셀 파일 경로 및 파일명
        file_path = fName

        # 엑셀 파일 불러오기
        workbook = load_workbook(file_path)

        # 첫 번째 시트 선택
        sheet = workbook.active

        # 'buyHistory' 컬럼의 마지막 행 인덱스 찾기
        last_row_index = sheet.max_row + 1

        # 새로운 sCode 추가
        sheet.cell(row=last_row_index, column=2, value=sCode)

        for idx, row_data in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            print('비교 - %s : %s' % (row_data[1], sCode))

        # 변경된 내용을 저장
        workbook.save(fName)

if __name__ == "__main__":
    Main()