import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook
import os

class Main():
    def __init__(self):
        sCode = '036640'
        self.buyHistory = ['036640', '050110', '096530', '158430']

        if sCode in self.buyHistory:
            self.buyHistory.remove(sCode)
            self.clearAndSaveExcel(self.buyHistory, 'buyHistory')

    def clearAndSaveExcel(self, data_list, file_name):
        # 디렉토리 생성
        directory = 'historyXlsx'
        if not os.path.exists(directory):
            os.makedirs(directory)

        # 엑셀 파일 경로
        file_path = os.path.join(directory, file_name + '.xlsx')

        # 기존 엑셀 파일이 있는 경우 내용 삭제 후 열기
        if os.path.exists(file_path):
            wb = load_workbook(file_path)
            ws = wb.active
            # 모든 셀 내용 삭제
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.value = None
            # 변경 사항 저장
            wb.save(file_path)
        else:
            # 새로운 Workbook 생성
            workbook = Workbook()
            # 첫 번째 시트 선택
            sheet = workbook.active
            # 컬럼명 설정
            sheet['A1'] = 'Index'
            sheet['B1'] = file_name
            workbook.save(file_path)

        # 새로운 데이터 추가
        workbook = load_workbook(file_path)
        sheet = workbook.active
        for index, data in enumerate(data_list):
            sheet[f'A{index + 2}'] = index
            sheet[f'B{index + 2}'] = data
        # 변경 사항 저장
        workbook.save(file_path)


if __name__ == "__main__":
    Main()
