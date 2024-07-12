import numpy as np
import sys
import os
import pandas as pd
import logging
from datetime import datetime
import pprint

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl import Workbook

from PyQt5.QAxContainer import *
from PyQt5.QtCore import *
from config.errorCode import *
from PyQt5.QtTest import *
from config.kiwoomType import *

class Kiwoom(QAxWidget):

    def __init__(self):
        super().__init__()

        today = datetime.today().date()
        ################## 로깅 설정 ##################
        logging.basicConfig(filename=f'logs/systemTrading_{today}.log', level=logging.DEBUG,
                            format="[ %(asctime)s | %(levelname)s ] %(message)s",
                            datefmt="%Y-%m-%d %H:%M:%S")
        self.logger = logging.getLogger()

        ################## 변수 모음 ##################
        self.realType = RealType()
        self.login_event_loop = None # 이벤트 루프 : 로그인
        self.detail_account_info_event_loop = QEventLoop() # 이벤트 루프 : 예수금 상세 정보 요청
        self.calculator_event_loop = QEventLoop()

        self.account_num = None # 보유 계좌번호
        self.use_money = 0
        self.use_money_persent = 0.5
        self.threshold = 3 # 매수 기준선과의 격차율

        # 스크린 번호
        self.screen_my_info = "2000"
        self.screen_calculation_stock = "4000"
        self.screen_real_stock = "5000" # 종목별 할당할 스크린 번호
        self.screen_meme_stock = "6000" # 종목별 할당할 주문용 스크린 번호
        self.screen_start_stop_real = "1000" # 개장, 폐장 여부에 대한 실시간 확인용 스크린 번호
        self.filePath = 'historyXlsx/'

        self.calcul_data = [] #종목별 일봉 데이터 저장 리스트
        self.account_stock_dict = {} # 보유하고 있는 종목 저장 딕셔너리
        self.not_account_stock_dick = {} # 미체결 주문 종목들의 집합
        self.portfolio_stock_dict = {} # 120일선 기준 필터링되어 txt파일에서 불러와 관심종목으로 저장할 딕셔너리
        self.jango_dict = {} # 당일 실시간 계좌 잔고 현황

        self.sumVol1 = {} # 거래량 증가 비교군 1
        self.sumVol2 = {} # 거래량 증가 비교군 2
        self.sumVol3 = {}  # 거래량 증가 비교군 3
        self.sumVol4 = {}  # 거래량 증가 비교군 4

        self.sumCrntPrce1 = {} # 호가 증가 비교군 1
        self.sumCrntPrce2 = {} # 호가 증가 비교군 2
        self.sumCrntPrce3 = {} # 호가 증가 비교군 3
        self.sumCrntPrce4 = {} # 호가 증가 비교군 4
        #self.tmpStockDict = {} # 종목별 dataframe을 한번만 수행하도록 담아두는 딕셔너리

        # 조건검색을 위한 변수모음
        self.gValueList = [] # 1번, 평균거래량을 구하기 위한 거래량 축적 리스트
        self.cci_low_data = {} # 종목별 cci를 구하기 위한 딕셔너리
        self.gCurrentList = [] # 2번, 5번 종가를 활용하기 위한 종가 리스트
        self.gCurrent = None # 3번 종가의 범의 확인을 위한 종가
        self.gStrength = None # 4번 체결강도
        self.gCapitalization = None # 6번 시가총액
        self.duringDays = 100
        self.gPass_success = False
        self.moving20_average_price = None

        # 부분 매도 / 추가 매수 / 같은 조건 반복 매수를 방지하기 위한 변수 모음
        self.buyHistory = [] # 매수 내역 리스트
        self.addBuyHistory = [] # 추가 매수 내역 리스트
        self.partialSalelist = [] # 부분 매도 리스트
        ################## 변수 모음 끝 ##################


        ################## 함수 실행 ##################
        self.get_ocx_instance()
        self.event_slots()
        self.real_event_slots()

        #self.txtToxlsx()
        self.signal_login_commConnect() # 로그인
        self.get_account_info() # 보유 계좌 조회
        self.detail_account_info() # 계좌에 대한 상세 정보 조회
        self.detail_account_mystock() # 보유 주식 조회
        self.not_concluded_account() # 실시간 미체결 종목 조회

        self.read_code02() #  120일선 기준 필터링되어 저장된 txt파일에서 self.portfolio_stock_dict로 종목 이관하는 함수 => 보유종목과 관종과의 괴리를 없애기 위해 엑셀 파일명으로 수정
        self.screen_number_setting() # 종목별 스크린번호 관리

        # 임시 실행 : 데이터를 받아와 조건에 맞는 종목을 선정하여 txt 파일로 저장한다. => 관심종목 추가 => 필요에 따라 수행하는 역할
        # self.file_delete()
        # self.calculator_fnc() # 주식 시장별 정보 조회

        ################## 함수 실행 끝 ##################


        ################## 장 시작/종료 및 실시간 정보 수집 시그널 ##################

        # 실시간 등록 - SetRealReg 요청에 대한 응답은 OnReceiveRealData
        self.logger.info("장시작/종료 실시간 요청")
        # print("real_event_slots() - 장시작/종료 실시간 요청")
        self.dynamicCall("SetRealReg(QString, QString, QString, QString", self.screen_start_stop_real, "", self.realType.REALTYPE['장시작시간']['장운영구분'],"0")

        # 종목 코드 등록
        for code in self.portfolio_stock_dict.keys():
            screen_num = self.portfolio_stock_dict[code]['스크린번호']
            fids = self.realType.REALTYPE['주식체결']['체결시간']

            self.dynamicCall("SetRealReg(QString, QString, QString, QString", screen_num, code, fids, "1")
            self.logger.info("[초기화정보세팅] 종목코드 등록 : %s, 스크린번호 : %s, fid 번호 : %s" % (code, screen_num, fids))
            print("[초기화정보세팅] 종목코드 등록 : %s, 스크린번호 : %s, fid 번호 : %s" % (code, screen_num, fids))

        ################## 장 시작/종료 및 실시간 정보 수집 시그널 끝 ##################

    def get_ocx_instance(self):
        self.setControl("KHOPENAPI.KHOpenAPICtrl.1")

    # 이벤트 집합소
    # 서버에 요청 후 리턴 받는 응답을 한 곳에서 받아 처리
    def event_slots(self):
        print("event_slots")
        #로그인 이벤트 응답 / errorCode.py에 코드별 상태값 정리 참고
        self.OnEventConnect.connect(self.login_slot) # 로그인 EventLoop
        #TR 이벤트 응답
        self.OnReceiveTrData.connect(self.trdata_slot)
        # msg 이벤트 응답
        self.OnReceiveMsg.connect(self.msg_slot)

    # 실시간 요청 컨드롤
    def real_event_slots(self):
        # print("real_event_slots")
        # 장시작/종료 실시간 응답 처리
        self.OnReceiveRealData.connect(self.realdata_slot)
        # 주문에 대한 이벤트 등록
        self.OnReceiveChejanData.connect(self.chejan_slot)

    def signal_login_commConnect(self):
        self.dynamicCall("CommConnect()") # 키움 로그인을 위한 메서드 이름과 사용 방법 : dynamicCall이라는 메서드를 이용하여 호출

        self.login_event_loop = QEventLoop() # 로그인 EventLoop 설정
        self.login_event_loop.exec_() # 로그인 EventLoop 시작

    def login_slot(self, errCode):
        self.logger.info(errors(errCode))
        # print(errors(errCode))

        self.login_event_loop.exit() # 로그인 EventLoop 끝
    # 내 정보 가져오기
    def get_account_info(self):
        account_list = self.dynamicCall("GetLoginInfo(String)", "ACCNO")
        self.logger.info('[초기정보세팅] 전체보유계좌 : %s' % account_list)
        print("[초기정보세팅] 전체보유계좌 : %s" % account_list)

        self.account_num = account_list.split(';')[1]
        self.logger.info('[초기정보세팅] 매매 계좌번호 : %s' % self.account_num)
        print('[초기정보세팅] 매매 계좌번호 : %s' % self.account_num)

    # 예수금 정보 가져오기
    def detail_account_info(self):
        # SetInputValue 서버 전송전 필요 데이터 입력
        # CommRqData 데이터 요청 실행
        # print("detail_account_info() => 예수금 요청")    # 키움 예수금 정보 요청하기 위한 메서드 이름과 사용 방법 : dynamicCall이라는 메서드를 이용하여 호출
        self.dynamicCall("SetInputValue(String, String)", "계좌번호", self.account_num)
        self.dynamicCall("SetInputValue(String, String)", "비밀번호", "9958")
        self.dynamicCall("SetInputValue(String, String)", "비밀번호입력매체구분", "00")
        self.dynamicCall("SetInputValue(String, String)", "조회구분", "2")
        self.dynamicCall("CommRqData(String, String, int, String)", "예수금상세현황요청", "OPW00001", "0", self.screen_my_info)

        self.detail_account_info_event_loop.exec_()

    def detail_account_mystock(self, sPrevNext="0"): # 계좌 평가 잔고 내역 요청
        self.dynamicCall("SetInputValue(String, String)", "계좌번호", self.account_num)
        self.dynamicCall("SetInputValue(String, String)", "비밀번호", "9958")
        self.dynamicCall("SetInputValue(String, String)", "비밀번호입력매체구분", "00")
        self.dynamicCall("SetInputValue(String, String)", "조회구분", "2")
        self.dynamicCall("CommRqData(String, String, int, String)", "계좌평가잔고내역요청", "OPW00018", sPrevNext, self.screen_my_info)

        if sPrevNext == "0" or sPrevNext == "":
            #self.detail_account_mystock_event_loop = QEventLoop()
            self.detail_account_info_event_loop.exec_()

    def not_concluded_account(self, sPrevNext="0"): # 미체결 종목 요청
        # print("not_concluded_account() => 실시간미체결종목요청")
        self.dynamicCall("SetInputValue(QString, QString)", "계좌번호", self.account_num)
        self.dynamicCall("SetInputValue(QString, QString)", "체결구분", "1")
        self.dynamicCall("SetInputValue(QString, QString)", "매매구분", "0")
        self.dynamicCall("CommRqData(String, String, int, String)", "실시간미체결종목요청", "opt10075", sPrevNext, self.screen_my_info)

        self.detail_account_info_event_loop.exec_()


    def trdata_slot(self, sScrNo, sRQName, sTrCode, sRecordName, sPrevNext): #
        print('trdata_slot')
        '''
        TR 요청에 대한 응답 처리
        sScrNo : 스크린번호
        sRQName : 요청명 - 개발자 마음대로
        sTrCode : 요청 ID TR코드
        sRecordName :  사용안함
        sPrevNext : 다음 페이지가 있는지
        return :
        '''
        # print("trdata_slot() => TR요청에 대한 응답")

        if sRQName == "예수금상세현황요청":
            deposit = self.dynamicCall("GetCommData(String, String, int, String)", sTrCode, sRQName, 0, "예수금")
            self.logger.info('[초기정보세팅] 예수금 : %s' % deposit)
            print("[초기정보세팅] 예수금 : %s" % deposit)

            ok_deposit = self.dynamicCall("GetCommData(String, String, int, String)", sTrCode, sRQName, 0, "출금가능금액")
            self.logger.info('[초기정보세팅] 출금가능금액 : %s' % ok_deposit)
            print("[초기정보세팅] 출금가능금액 : %s" % ok_deposit)

            ##############################################################
            # 매수 가능 금액 조절 - 몰빵하는 것을 시스템 적으로 조정하는 기능 (차후 조정하면서 매매할 수 있음)
            ##############################################################
            self.use_money = int(deposit) * self.use_money_persent
            # self.use_money = self.use_money / 4
            ##############################################################

            self.detail_account_info_event_loop.exit() # 예수금 상세 현황 요청 이벤트 루트 종료

        elif sRQName == "계좌평가잔고내역요청":
            # print("[trdata_slot] sRQName :: 계좌평가잔고내역요청")
            # 총 매입 금액 / 총 수익률
            total_buy_money = self.dynamicCall("GetCommData(String, String, int, String)", sTrCode, sRQName, 0, "총매입금액")
            total_buy_money_result = int(total_buy_money)

            self.logger.info('[초기정보세팅] 총매입금액 : %s' % total_buy_money_result)
            print("[초기정보세팅] 총매입금액 : %s" % total_buy_money_result)

            total_progit_loss_rate = self.dynamicCall("GetCommData(String, String, int, String)", sTrCode, sRQName, 0, "총수익률(%)")
            total_progit_loss_rate_result = float(total_progit_loss_rate)
            self.logger.info('[초기정보세팅] 총수익률(%%) : %s' % total_progit_loss_rate_result)
            print("[초기정보세팅] 총수익률(%%) : %s" % total_progit_loss_rate_result)

            # 보유 종목 개수
            rows = self.dynamicCall("GetRepeatCnt(QString, QString)", sTrCode, sRQName)
            self.logger.info('[초기정보세팅] 보유주 개수 : %s' % rows)
            print("[초기정보세팅] 보유주 개수 : %s" % rows)
            cnt = 0
            for i in range(rows):
                code = self.dynamicCall("GetCommData(QString, QString, int, QString)", sTrCode, sRQName, i,"종목번호")
                code = code.strip()[1:]

                code_nm = self.dynamicCall("GetCommData(QString, QString, int, QString)", sTrCode, sRQName, i,"종목명")
                stock_quantity = self.dynamicCall("GetCommData(QString, QString, int, QString)", sTrCode, sRQName, i,"보유수량")
                buy_price = self.dynamicCall("GetCommData(QString, QString, int, QString)", sTrCode, sRQName, i,"매입가")
                learn_rate = self.dynamicCall("GetCommData(QString, QString, int, QString)", sTrCode, sRQName, i,"수익률(%)")
                current_price = self.dynamicCall("GetCommData(QString, QString, int, QString)", sTrCode, sRQName, i,"현재가")
                total_chegual_price = self.dynamicCall("GetCommData(QString, QString, int, QString)", sTrCode, sRQName, i,"매입금액")
                possible_quantity = self.dynamicCall("GetCommData(QString, QString, int, QString)", sTrCode, sRQName, i,"매매가능수량")

                if code in self.account_stock_dict:
                    pass
                else:
                    self.account_stock_dict.update({code:{}})

                code_nm = code_nm.strip()
                stock_quantity = int(stock_quantity.strip())
                buy_price = int(buy_price.strip())
                learn_rate = float(learn_rate.strip())
                current_price = int(current_price.strip())
                total_chegual_price = int(total_chegual_price.strip())
                possible_quantity = int(possible_quantity.strip())

                self.account_stock_dict[code].update({"종목명": code_nm})
                self.account_stock_dict[code].update({"보유수량": stock_quantity})
                self.account_stock_dict[code].update({"매입가": buy_price})
                self.account_stock_dict[code].update({"수익률(%)": learn_rate})
                self.account_stock_dict[code].update({"현재가": current_price})
                self.account_stock_dict[code].update({"매입금액": total_chegual_price})
                self.account_stock_dict[code].update({"매매가능수량": possible_quantity})

                cnt += 1
            self.logger.info('[초기정보세팅] 매매계좌에 가지고 있는 종목 : %s' % self.account_stock_dict)
            print("[초기정보세팅] 매매계좌에 가지고 있는 종목 : %s " % self.account_stock_dict)
            if sPrevNext == "":
                sPrevNext = "0"

            if sPrevNext == "2":
                self.detail_account_mystock(sPrevNext="2")
            else:
                self.detail_account_info_event_loop.exit()

        elif sRQName == "실시간미체결종목요청":
            rows = self.dynamicCall("GetRepeatCnt(QString, QString)", sTrCode, sRQName)
            for i in range(rows):
                code = self.dynamicCall("GetCommData(QString, QString, int, QString)", sTrCode, sRQName, i,"종목코드")
                code_nm = self.dynamicCall("GetCommData(QString, QString, int, QString)", sTrCode, sRQName, i,"종목명")
                order_no = self.dynamicCall("GetCommData(QString, QString, int, QString)", sTrCode, sRQName, i,"주문번호")
                order_status = self.dynamicCall("GetCommData(QString, QString, int, QString)", sTrCode, sRQName, i,"주문상태")
                order_quantity = self.dynamicCall("GetCommData(QString, QString, int, QString)", sTrCode, sRQName, i,"주문수량")
                order_price = self.dynamicCall("GetCommData(QString, QString, int, QString)", sTrCode, sRQName, i,"주문가격")
                order_gubun = self.dynamicCall("GetCommData(QString, QString, int, QString)", sTrCode, sRQName, i,"주문구분")
                not_quantity = self.dynamicCall("GetCommData(QString, QString, int, QString)", sTrCode, sRQName, i,"미체결수량")
                ok_quantity = self.dynamicCall("GetCommData(QString, QString, int, QString)", sTrCode, sRQName, i,"체결량")

                code = code.strip()
                code_nm = code_nm.strip()
                order_no = int(order_no.strip())
                order_status = order_status.strip()
                order_quantity = int(order_quantity.strip())
                order_price = int(order_price.strip())
                order_gubun = order_gubun.strip().lstrip("+").lstrip("-")
                not_quantity = int(not_quantity.strip())
                ok_quantity = int(ok_quantity.strip())

                if order_no in self.not_account_stock_dick:
                    pass
                else:
                    self.not_account_stock_dick[order_no] = {}

                nasd = self.not_account_stock_dick[order_no]

                nasd.update({"종목코드": code})
                nasd.update({"종목명": code_nm})
                nasd.update({"주문번호": order_no})
                nasd.update({"주문상태": order_status})
                nasd.update({"주문수량": order_quantity})
                nasd.update({"주문가격": order_price})
                nasd.update({"주문구분": order_gubun})
                nasd.update({"미체결수량": not_quantity})
                nasd.update({"체결량": ok_quantity})

                #self.logger.info('[실시간] 미체결 종목 : %s' % self.not_account_stock_dick[order_no])
                print("[실시간] 미체결 종목 : %s " % self.not_account_stock_dick[order_no])

            self.detail_account_info_event_loop.exit()

        elif sRQName == "주식일봉차트조회":
            print('주식일봉차트조회')
            code = self.dynamicCall("GetCommData(QString, QString, int, QString)", sTrCode, sRQName, 0,"종목코드")
            code = code.strip()
            # print("[info] %s 종목" % code)

            # 해당 종목의 거래일 수 조회
            cnt = self.dynamicCall("GetRepeatCnt(QString, QString)", sTrCode, sRQName)

            ########################################################################################################
            # 해당 종목의 모든 거래일 데이터를 받아와서 self.calcul_data에 저장
            ########################################################################################################
            for i in range(cnt):
                data = []

                current_price = self.dynamicCall("GetCommData(QString, QString, int, QString)", sTrCode, sRQName, i,"현재가")
                value = self.dynamicCall("GetCommData(QString, QString, int, QString)", sTrCode, sRQName, i,"거래량")
                trading_value = self.dynamicCall("GetCommData(QString, QString, int, QString)", sTrCode, sRQName, i,"거래대금")
                date = self.dynamicCall("GetCommData(QString, QString, int, QString)", sTrCode, sRQName, i,"일자")
                start_price = self.dynamicCall("GetCommData(QString, QString, int, QString)", sTrCode, sRQName, i,"시가")
                high_price = self.dynamicCall("GetCommData(QString, QString, int, QString)", sTrCode, sRQName, i,"고가")
                low_price = self.dynamicCall("GetCommData(QString, QString, int, QString)", sTrCode, sRQName, i,"저가")

                data.append("")
                data.append(current_price.strip())
                data.append(value.strip())
                data.append(trading_value.strip())
                data.append(date.strip())
                data.append(start_price.strip())
                data.append(high_price.strip())
                data.append(low_price.strip())
                data.append("")

                self.calcul_data.append(data.copy())

                # ['', '316', '376190', '119', '20190326', '310', '323', '310', '']
                #     현재가 ,  거래량  ,거래대금,    일자   ,  시가,   고가,   저가
                # 0일부터 총 조회 일수까지 루프

            if sPrevNext == "2":
                self.day_kiwoom_db(code=code, sPrevNext=sPrevNext)
            else: # sPrevNext가 0 또는 Null일 경우 - 조회된 거래일 데이터 수가 600개 미만일 경우

                self.calcul_data.clear()
                self.calculator_event_loop.exit()

    def realdata_slot(self, sCode, sRealType, sRealData):
        if sRealType == "장시작시간":
            fid = self.realType.REALTYPE[sRealType]['장운영구분']
            value = self.dynamicCall("GetCommRealData(QString, int)", sCode, fid)

            if value == '0':
                print("[info] 장시작 전")
            elif value == '3':
                print("[info] 장시작")
            elif value == '2':
                print("[info] 장 종료, 동시호가로 넘어감")
            elif value == '4':
                print("[info] 15시 30분 장종료")

                for code in self.portfolio_stock_dict.keys():
                    self.dynamicCall("SetRealRemove(QString, QString)", self.portfolio_stock_dict[code]['스크린번호'], code)
                    code_nm = self.dynamicCall("GetMasterCodeName(QString)", code)  # 종목코드 1개의 종목 한글명을 반환한다.
                    #if code not in self.account_stock_dict.keys():
                    #    self.file_delete(code, code_nm)

                self.logger.info('[장종료] self.buyHistory : %s ' % self.buyHistory)
                self.logger.info('[장종료]self.addBuyHistory : %s ' % self.addBuyHistory)
                self.logger.info('[장종료] self.partialSalelist : %s ' % self.partialSalelist)

                QTest.qWait(5000)
                #self.calculator_fnc()

                sys.exit()

        elif sRealType == "주식체결":  ##################
            # 체결시간 => HHMMSS
            a = self.dynamicCall("GetCommRealData(QString, int)", sCode, self.realType.REALTYPE[sRealType]['체결시간'])
            b = self.dynamicCall("GetCommRealData(QString, int)", sCode, self.realType.REALTYPE[sRealType]['현재가'])
            b = abs(int(b))  # abs : 절대값
            c = self.dynamicCall("GetCommRealData(QString, int)", sCode, self.realType.REALTYPE[sRealType]['전일대비'])
            c = abs(int(c))  # abs : 절대값
            d = self.dynamicCall("GetCommRealData(QString, int)", sCode, self.realType.REALTYPE[sRealType]['등락율'])
            d = float(d)
            e = self.dynamicCall("GetCommRealData(QString, int)", sCode, self.realType.REALTYPE[sRealType]['(최우선)매도호가'])
            e = abs(int(e))
            f = self.dynamicCall("GetCommRealData(QString, int)", sCode, self.realType.REALTYPE[sRealType]['(최우선)매수호가'])
            f = abs(int(f))
            g = self.dynamicCall("GetCommRealData(QString, int)", sCode, self.realType.REALTYPE[sRealType]['거래량'])
            g = abs(int(g))
            h = self.dynamicCall("GetCommRealData(QString, int)", sCode, self.realType.REALTYPE[sRealType]['누적거래량'])
            h = abs(int(h))
            i = self.dynamicCall("GetCommRealData(QString, int)", sCode, self.realType.REALTYPE[sRealType]['고가'])
            i = abs(int(i))
            j = self.dynamicCall("GetCommRealData(QString, int)", sCode, self.realType.REALTYPE[sRealType]['시가'])
            j = abs(int(j))
            k = self.dynamicCall("GetCommRealData(QString, int)", sCode, self.realType.REALTYPE[sRealType]['저가'])
            k = abs(int(k))
            m = self.dynamicCall("GetCommRealData(QString, int)", sCode, self.realType.REALTYPE[sRealType]['시가총액(억)'])
            m = abs(int(m))
            o = self.dynamicCall("GetCommRealData(QString, int)", sCode, self.realType.REALTYPE[sRealType]['체결강도'])
            o = abs(int(float(o)))

            code_nm = self.dynamicCall("GetMasterCodeName(QString)", sCode)  # 종목코드 1개의 종목 한글명을 반환한다.

            benchmark = 20

            # 체결가 증가 비교군
            if sCode not in self.sumCrntPrce1:
                self.sumCrntPrce1.update({sCode:{'crntprce':b, 'count':1}})
                self.sumCrntPrce2.update({sCode:{'crntprce':b, 'count':1}})
                self.sumCrntPrce3.update({sCode:{'crntprce':b, 'count':1}})
                self.sumCrntPrce4.update({sCode:{'crntprce':b, 'count':1}})
            else:
                if self.sumCrntPrce1[sCode]['count'] >= benchmark:
                    pass
                else:
                    self.sumCrntPrce1[sCode]['count'] += 1
                    self.sumCrntPrce1[sCode]['crntprce'] += b

            if self.sumCrntPrce2[sCode]['count'] >= benchmark:
                pass
                #self.sumVol1[sCode] = self.sumVol2[sCode]
                #del self.sumVol2[sCode]
            elif self.sumCrntPrce1[sCode]['count'] >= benchmark:
                self.sumCrntPrce2[sCode]['count'] += 1
                self.sumCrntPrce2[sCode]['crntprce'] += b

            if self.sumCrntPrce3[sCode]['count'] >= benchmark:
                pass
                #self.sumVol1[sCode] = self.sumVol2[sCode]
                #del self.sumVol2[sCode]
            elif self.sumCrntPrce1[sCode]['count'] >= benchmark and self.sumCrntPrce2[sCode]['count'] >= benchmark:
                self.sumCrntPrce3[sCode]['count'] += 1
                self.sumCrntPrce3[sCode]['crntprce'] += b

            if self.sumCrntPrce4[sCode]['count'] >= benchmark:
                self.sumCrntPrce1[sCode] = self.sumCrntPrce2[sCode]
                self.sumCrntPrce2[sCode] = self.sumCrntPrce3[sCode]
                self.sumCrntPrce3[sCode] = self.sumCrntPrce4[sCode]
                self.sumCrntPrce4[sCode] = {'crntprce': 1, 'count': 1}
            elif self.sumCrntPrce1[sCode]['count'] >= benchmark and self.sumCrntPrce2[sCode]['count'] >= benchmark and self.sumCrntPrce3[sCode]['count'] >= benchmark:
                self.sumCrntPrce4[sCode]['count'] += 1
                self.sumCrntPrce4[sCode]['crntprce'] += b

            # 실시간 분봉 데이터 받는 코드 짜기가 귀찮아서 임시 대체 해봄
            if sCode not in self.sumVol1:
                self.sumVol1.update({sCode:{'vol':g, 'count':1}})
                self.sumVol2.update({sCode:{'vol':g, 'count':1}})
                self.sumVol3.update({sCode:{'vol':g, 'count':1}})
                self.sumVol4.update({sCode:{'vol':g, 'count':1}})
            else:
                if self.sumVol1[sCode]['count'] >= benchmark:
                    pass
                else:
                    self.sumVol1[sCode]['count'] += 1
                    self.sumVol1[sCode]['vol'] += g

            if self.sumVol2[sCode]['count'] >= benchmark:
                pass
                #self.sumVol1[sCode] = self.sumVol2[sCode]
                #del self.sumVol2[sCode]
            elif self.sumVol1[sCode]['count'] >= benchmark:
                self.sumVol2[sCode]['count'] += 1
                self.sumVol2[sCode]['vol'] += g

            if self.sumVol3[sCode]['count'] >= benchmark:
                pass
                #self.sumVol1[sCode] = self.sumVol2[sCode]
                #del self.sumVol2[sCode]
            elif self.sumVol1[sCode]['count'] >= benchmark and self.sumVol2[sCode]['count'] >= benchmark:
                self.sumVol3[sCode]['count'] += 1
                self.sumVol3[sCode]['vol'] += g

            if self.sumVol4[sCode]['count'] >= benchmark:
                self.sumVol1[sCode] = self.sumVol2[sCode]
                self.sumVol2[sCode] = self.sumVol3[sCode]
                self.sumVol3[sCode] = self.sumVol4[sCode]
                self.sumVol4[sCode] = {'vol': 1, 'count': 1}
            elif self.sumVol1[sCode]['count'] >= benchmark and self.sumVol2[sCode]['count'] >= benchmark and self.sumVol3[sCode]['count'] >= benchmark:
                self.sumVol4[sCode]['count'] += 1
                self.sumVol4[sCode]['vol'] += g

            if sCode not in self.portfolio_stock_dict:  # 체결된 종목이 나의 관심 종목에 저장되어 있지 않다면
                print("☆☆☆☆☆☆☆☆☆☆☆☆☆☆☆☆☆☆☆☆☆☆☆☆☆ self.portfolio_stock_dict에 등록 되어 있지 않습니다. : %s (%s)" % (code_nm, sCode))
                self.portfolio_stock_dict.update({sCode: {}})  # 추가한다.
                #self.getLowData(sCode)
            if sCode in self.portfolio_stock_dict:
                self.portfolio_stock_dict[sCode].update({'체결시간': a})
                self.portfolio_stock_dict[sCode].update({'현재가': b})
                self.portfolio_stock_dict[sCode].update({'전일대비': c})
                self.portfolio_stock_dict[sCode].update({'등락율': d})
                self.portfolio_stock_dict[sCode].update({'(최우선)매도호가': e})
                self.portfolio_stock_dict[sCode].update({'(최우선)매수호가': f})
                self.portfolio_stock_dict[sCode].update({'거래량': g})
                self.portfolio_stock_dict[sCode].update({'누적거래량': h})
                self.portfolio_stock_dict[sCode].update({'고가': i})
                self.portfolio_stock_dict[sCode].update({'시가': j})
                self.portfolio_stock_dict[sCode].update({'저가': k})

            ma20_price = self.getMaPrice(sCode, 20)
            print('ma20_price : %s' % ma20_price)
            ma20_ratio = ((b - int(ma20_price)) / int(ma20_price))

            print('-----------------------------------------------------------------------------------------------------------------------------')
            print('[ANALYZE] (현재시간 : %s ) 종목명 : %s(%s)' % (datetime.now(), code_nm, sCode))
            print('[ANALYZE] CCI 신호검색 결과 => %s' % self.getCCI(i, k, b, sCode, 50))
            print('[ANALYZE] 현재가(%s) >  20평균가(%s)' % (b, ma20_price))
            print('[ANALYZE] 주가누적1(%s/%s), 주가누적2(%s/%s), 주가누적3(%s/%s)' % (
                self.sumCrntPrce1[sCode]['count'], self.sumCrntPrce1[sCode]['crntprce'],
                self.sumCrntPrce2[sCode]['count'], self.sumCrntPrce2[sCode]['crntprce'],
                self.sumCrntPrce3[sCode]['count'], self.sumCrntPrce3[sCode]['crntprce']))
            print("[ANALYZE] 0.5 <= 이격률(%s) < 3 |" % (ma20_ratio * 100))

            self.logger.info('-----------------------------------------------------------------------------------------------------------------------------')
            self.logger.info('[ANALYZE] (현재시간 : %s ) 종목명 : %s(%s)' % (datetime.now(), code_nm, sCode))
            self.logger.info('[ANALYZE] CCI 신호검색 결과 => %s' % self.getCCI(i, k, b, sCode, 50))
            self.logger.info('[ANALYZE] 현재가(%s) >  20평균가(%s)' % (b, ma20_price))
            self.logger.info('[ANALYZE] 현재가1(%s/%s), 현재가2(%s/%s), 현재가3(%s/%s)' % (
                    self.sumCrntPrce1[sCode]['count'], self.sumCrntPrce1[sCode]['crntprce'],
                    self.sumCrntPrce2[sCode]['count'],self.sumCrntPrce2[sCode]['crntprce'],
                    self.sumCrntPrce3[sCode]['count'],self.sumCrntPrce3[sCode]['crntprce']))
            self.logger.info("[ANALYZE] 0.5 <= 이격률(%s) < 3 |" % (ma20_ratio * 100))

            # 1. 계좌 잔고 평가 내역에 있고 당일 실시간 잔고에 없음  - 61강 2분 15초 이후
            if sCode in self.account_stock_dict.keys() and sCode not in self.jango_dict.keys():

                asd = self.account_stock_dict[sCode]  # #계좌 평가 잔고 내역 리스트
                meme_rate = (b - asd['매입가']) / asd['매입가'] * 100

                # 20일평균 기준 체결가의 비율만큼
                # 체결가를 기준으로 상기 비율만큼을 반영한 수익 목표가액
                increased_price = asd['매입가'] + (asd['매입가'] * ma20_ratio)
                print('[ANALYZE] [보유주] 수익목표가액 : %s 이상' % meme_rate)

                # self.logger.info("[틱발생] (보유주) 매매가능수량 : %s | meme_rate : %s" % (asd['매매가능수량'], meme_rate))
                # print("[틱발생] (보유주) 매매가능수량 : %s | meme_rate : %s" % (asd['매매가능수량'], meme_rate))
                self.logger.info('[DEBUG] 보유주 수익목표가액 : %s 이상' % meme_rate)
                if asd['매매가능수량'] > 0:
                    if meme_rate >= 10:

                        print('[보유주 - 익절매도조건] 현재가(%s) >= 상승비율반영가(%s))' % (b, increased_price))
                        self.logger.info('[보유주 - 익절매도조건] 현재가(%s) >= 상승비율반영가(%s))' % (b, increased_price))
                        self.logger.info("[보유주 - 익절매도] %s(%s) | 매도 수량 : %s | 체결단가 : %s | 매입가 : %s | 20평균 : %s | 상승비율반영가 : %s"
                                         % (code_nm, sCode, asd['매매가능수량']*0.5, b, asd['매입가'], ma20_price, increased_price))

                        order_success = self.dynamicCall(
                            "SendOrder(QString, QString, QString, int, QString, int, int, QString, QString",
                            ["신규매도", self.portfolio_stock_dict[sCode]['주문용스크린번호'], self.account_num, 2,
                             sCode, asd['매매가능수량']*0.5, 0, self.realType.SENDTYPE['거래구분']['시장가'], ""])

                        if sCode not in self.partialSalelist:
                            self.partialSalelist.append(str(sCode))
                            self.clearAndSaveExcel(self.partialSalelist, 'partialSalelist')

                        self.historyListLogging(code_nm, sCode, '보유주 - 익절매도')

                    elif b < ma20_price:

                        print('[보유주 - 손절매도조건] 현재가(%s) < 20일평균(%s))' % (b, ma20_price))
                        self.logger.info('[보유주 - 손절매도조건] 현재가(%s) < 20일평균(%s))' % (b, ma20_price))

                        self.logger.info("[보유주 - 손절매도] %s(%s) | 매도 수량 : %s | 체결단가 : %s | 매입가 : %s | 20평균 : %s"
                                         % (code_nm, sCode, asd['매매가능수량'], b, asd['매입가'], ma20_price))
                        print("[보유주 - 손절매도] %s(%s) | 매도 수량 : %s | 체결단가 : %s | 매입가 : %s | 20평균 : %s"
                              % (code_nm, sCode, asd['매매가능수량'], b, asd['매입가'], ma20_price))

                        order_success = self.dynamicCall(
                            "SendOrder(QString, QString, QString, int, QString, int, int, QString, QString",
                            ["신규매도", self.portfolio_stock_dict[sCode]['주문용스크린번호'], self.account_num, 2,
                             sCode, asd['매매가능수량'], 0, self.realType.SENDTYPE['거래구분']['시장가'], ""])

                        if sCode in self.buyHistory:
                            self.buyHistory.remove(str(sCode))
                            self.xlsxDelete(str(sCode),  self.filePath+'buyHistory.xlsx')

                        if sCode in self.addBuyHistory:
                            self.addBuyHistory.remove(str(sCode))
                            self.xlsxDelete(str(sCode),  self.filePath+'addBuyHistory.xlsx')

                        if sCode in self.partialSalelist:
                            self.partialSalelist.remove(str(sCode))
                            self.xlsxDelete(str(sCode),  self.filePath+'partialSalelist.xlsx')

                        self.historyListLogging(code_nm, str(sCode), '보유주 - 손절매도')
                    else:
                        order_success = None


                    if order_success == 0:
                        self.logger.info("[info] 매도 주문 전달 성공")
                        print("[info] 매도 주문 전달 성공")
                        del self.account_stock_dict[sCode]
                    elif order_success == None:
                        pass
                        # self.logger.info("[info] 매도주문 없음")
                    else:
                        self.logger.info("[info] 매도주문 전달 실패")
                        # print("[info] 매도주문 전달 실패")
            # 1. 계좌 잔고 평가 내역에 없고 당일 실시간 잔고에 있음
            elif sCode in self.jango_dict.keys():
                jd = self.jango_dict[sCode]
                meme_rate = (b - jd['매입단가']) / jd['매입단가'] * 100

                # 20일평균 기준 체결가의 비율만큼
                # 체결가를 기준으로 상기 비율만큼을 반영한 수익 목표가액
                increased_price = jd['매입단가'] + (jd['매입단가'] * ma20_ratio)
                print('[ANALYZE] [당일주] 수익목표가액 : %s 이상' % meme_rate)

                if jd['주문가능수량'] > 0:
                    if meme_rate >= 3:
                        order_success = self.dynamicCall(
                            "SendOrder(QString, QString, QString, int, QString, int, int, QString, QString",
                            ["신규매도", self.portfolio_stock_dict[sCode]['주문용스크린번호'], self.account_num, 2,
                             sCode, jd['주문가능수량']*0.5, 0, self.realType.SENDTYPE['거래구분']['시장가'], ""])

                        if sCode not in self.partialSalelist:
                            self.partialSalelist.append(str(sCode))
                            self.clearAndSaveExcel(self.partialSalelist, 'partialSalelist')

                        self.historyListLogging(code_nm, sCode, '잔고주 - 익절매도')

                    elif b < ma20_price:

                        print('[당일주 - 손절매도조건] 현재가(%s) < 20일평균(%s)' % (b, ma20_price))
                        self.logger.info('[당일주 - 손절매도조건] 현재가(%s) < 20일평균(%s)' % (b, ma20_price))

                        order_success = self.dynamicCall(
                            "SendOrder(QString, QString, QString, int, QString, int, int, QString, QString",
                            ["신규매도", self.portfolio_stock_dict[sCode]['주문용스크린번호'], self.account_num, 2,
                             sCode, jd['주문가능수량'], 0, self.realType.SENDTYPE['거래구분']['시장가'], ""])

                        if sCode in self.buyHistory:
                            self.buyHistory.remove(str(sCode))
                            self.xlsxDelete(str(sCode),  self.filePath+'buyHistory.xlsx')

                        if sCode in self.addBuyHistory:
                            self.addBuyHistory.remove(str(sCode))
                            self.xlsxDelete(str(sCode),  self.filePath+'addBuyHistory.xlsx')

                        if sCode in self.partialSalelist:
                            self.partialSalelist.remove(str(sCode))
                            self.xlsxDelete(str(sCode),  self.filePath+'partialSalelist.xlsx')

                        self.historyListLogging(code_nm, str(sCode), '손절매도')
                    else:
                        order_success = None

                    if order_success == 0:
                        self.logger.info("[info] 매도주문 전달 성공")
                        print("[info] 매도주문 전달 성공")
                    elif order_success == None:
                        pass
                        #self.logger.info("[info] 매도주문 없음")
                    else:
                        self.logger.info("[info] 매도주문 전달 실패")
                        #print("[info] 매도주문 전달 실패")
            elif b > ma20_price and self.sumCrntPrce1[sCode]['crntprce'] < self.sumCrntPrce2[sCode]['crntprce'] and self.sumCrntPrce2[sCode]['crntprce'] < self.sumCrntPrce3[sCode]['crntprce']:
                if e != 0:
                    result = (self.use_money * 0.1) / e
                else:
                    result = 0

                quantity = int(result)
                order_success = None

                if quantity > 0:
                    print("[매수조건] 매수가능수량 : %s" % quantity)
                    if self.getCCI(i, k, b, sCode, 50) == True:
                        print('[매수조건] CCI 부합')

                        if sCode in self.buyHistory or sCode in self.addBuyHistory:
                            pass
                        elif 0.5 <= ma20_ratio * 100 < 3:
                            print("[매수조건] 0.5 <= 이격률(%s) < 3 |" % (ma20_ratio * 100))
                            self.logger.info("[매수조건] 0.5 <= 이격률(%s) < 3 |" % (ma20_ratio * 100))

                            order_success = self.dynamicCall(
                                "SendOrder(QString, QString, QString, int, QString, int, int, QString, QString",
                               ["신규매수", self.portfolio_stock_dict[sCode]['주문용스크린번호'], self.account_num, 1, sCode, quantity,
                                 e, self.realType.SENDTYPE['거래구분']['지정가'], ""])

                            self.buyHistory.append(str(sCode))
                            self.xlsxAdd(str(sCode), self.filePath+'buyHistory.xlsx')

                            self.historyListLogging(code_nm, sCode, '신규매수')
                        else:
                           self.logger.info("[info] 이격도 조건 불합격 : 0.5 < %s < 2" % (ma20_ratio * 100))
                           print("[info]  이격도 조건 불합격 : 0.5 < %s < 2" % (ma20_ratio * 100))

                        if order_success == 0:
                           self.logger.info("[info] 매수주문 전달 성공")
                           print("[info] 매수주문 전달 성공")
                        else:
                            pass
                            #self.logger.info("[info] 매수주문 전달 실패")
                            #print("[info] 매수주문 전달 실패")
                    else:
                       self.logger.info("[info] CCI신호 불합격")
                       print("[info] CCI신호 불합격")

                else:
                    # 종목명/종목코드/현재가)
                   self.logger.info("[info] 총알 부족으로 매수 주문 불가 - 종목 : %s(%s) | 현재가 : %s | 주문가능금액 : %s(%s)" % (code_nm, sCode, b, self.use_money * 0.1, quantity))
                   print("[info] 총알 부족으로 매수 주문 불가 - 종목 : %s(%s) | 현재가 : %s | 주문가능금액 : %s(%s)" % (code_nm, sCode, b, self.use_money * 0.1, quantity))

            # 미체결 종목
            # 루프 도중 미체결 리스트의 수량이 변경되어 오류가 발생하는 현상을 방지하기 위해 임의의 변수에 할당하여 루프 실행
            not_meme_list = list(self.not_account_stock_dick)
            for order_num in not_meme_list:
                code = self.not_account_stock_dick[order_num]["종목코드"]
                meme_price = self.not_account_stock_dick[order_num]["주문가격"]
                not_quantity = self.not_account_stock_dick[order_num]["미체결수량"]
                order_gubun = self.not_account_stock_dick[order_num]['주문구분']

                # 매수 취소
                if order_gubun == "신규매수" and not_quantity > 0 and e > meme_price:
                    self.logger.info("매수취소한다.", sCode)
                    # print("%s %s", ("매수취소한다.", sCode))
                    order_success = self.dynamicCall(
                        "SendOrder(QString, QString, QString, int, QString, int, int, QString, QString",
                        ["매수취소", self.portfolio_stock_dict[sCode]['주문용스크린번호'], self.account_num, 3, code, 0, 0,
                         self.realType.SENDTYPE['거래구분']['지정가'], order_num])

                    if order_success == 0:
                        self.logger.info("매수 취소 전달 성공.")
                        # print("[info] 매수 취소 전달 성공")
                    else:
                        self.logger.info("매수 취소 전달 실패.")
                        # print("[info] 매수 취소 전달 실패")


                elif not_quantity == 0:
                    del self.not_account_stock_dick[order_num]



    def chejan_slot(self, sGubun, nItemCnt, sFidLit):
        #print("chejan_slot() => 나의 매매 주문 요청에 대한 응답")
        #self.logger.info("[chejan_slot - 매매 요청에 대한 주문 응답 이벤트]")
        intGubun = int(sGubun)
        if intGubun == 0:
            self.logger.info("[chejan_slot] 매매 요청에 대한 주문 구분 : 미체결")
            account_num = self.dynamicCall("GetChejanData(int)", self.realType.REALTYPE['주문체결']['계좌번호'])
            sCode = self.dynamicCall("GetChejanData(int)", self.realType.REALTYPE['주문체결']['종목코드'])[1:]
            stock_name = self.dynamicCall("GetChejanData(int)", self.realType.REALTYPE['주문체결']['종목명'])
            stock_name = stock_name.strip()
            origin_number = self.dynamicCall("GetChejanData(int)", self.realType.REALTYPE['주문체결']['원주문번호'])
            order_number = self.dynamicCall("GetChejanData(int)", self.realType.REALTYPE['주문체결']['주문번호'])
            order_status = self.dynamicCall("GetChejanData(int)", self.realType.REALTYPE['주문체결']['주문상태'])
            order_quan = self.dynamicCall("GetChejanData(int)", self.realType.REALTYPE['주문체결']['주문수량'])
            order_quan = int(order_quan)
            order_price = self.dynamicCall("GetChejanData(int)", self.realType.REALTYPE['주문체결']['주문가격'])
            order_price = int(order_price)
            not_chegual_qaun = self.dynamicCall("GetChejanData(int)", self.realType.REALTYPE['주문체결']['미체결수량'])
            not_chegual_qaun = int(not_chegual_qaun)
            order_gubun = self.dynamicCall("GetChejanData(int)", self.realType.REALTYPE['주문체결']['주문구분'])
            order_gubun = order_gubun.strip().lstrip('+').lstrip('-')
            chegual_time_str = self.dynamicCall("GetChejanData(int)", self.realType.REALTYPE['주문체결']['주문/체결시간'])
            chegual_price = self.dynamicCall("GetChejanData(int)", self.realType.REALTYPE['주문체결']['체결가'])
            if chegual_price == '':
                chegual_price = 0
            else:
                chegual_price = int(chegual_price)

            chegual_quantity = self.dynamicCall("GetChejanData(int)", self.realType.REALTYPE['주문체결']['체결량'])
            if chegual_quantity == '':
                chegual_quantity = 0
            else:
                chegual_quantity = int(chegual_quantity)

            current_price = self.dynamicCall("GetChejanData(int)", self.realType.REALTYPE['주문체결']['현재가'])
            current_price = abs(int(current_price))

            first_sell_price = self.dynamicCall("GetChejanData(int)", self.realType.REALTYPE['주문체결']['(최우선)매도호가'])
            first_sell_price = abs(int(first_sell_price))

            first_buy_price = self.dynamicCall("GetChejanData(int)", self.realType.REALTYPE['주문체결']['(최우선)매수호가'])
            first_buy_price = abs(int(first_buy_price))

            # 새로 들어온  주문이면 주문번호 할당
            if order_number not in self.not_account_stock_dick.keys():
                self.not_account_stock_dick.update({order_number: {}})

            self.not_account_stock_dick[order_number].update({"종목코드": sCode})
            self.not_account_stock_dick[order_number].update({"주문번호": order_number})
            self.not_account_stock_dick[order_number].update({"종목명": stock_name})
            self.not_account_stock_dick[order_number].update({"주문상태": order_status})
            self.not_account_stock_dick[order_number].update({"주문수량": order_quan})
            self.not_account_stock_dick[order_number].update({"주문가격": order_price})
            self.not_account_stock_dick[order_number].update({"미체결수량": not_chegual_qaun})
            self.not_account_stock_dick[order_number].update({"원주문번호": origin_number})
            self.not_account_stock_dick[order_number].update({"주문구분": order_gubun})
            self.not_account_stock_dick[order_number].update({"주문/체결시간": chegual_time_str})
            self.not_account_stock_dick[order_number].update({"체결가": chegual_price})
            self.not_account_stock_dick[order_number].update({"체결량": chegual_quantity})
            self.not_account_stock_dick[order_number].update({"현재가": current_price})
            self.not_account_stock_dick[order_number].update({"(최우선)매도호가": first_sell_price})
            self.not_account_stock_dick[order_number].update({"(최우선)매수호가": first_buy_price})

        elif intGubun == 1:
            self.logger.info("[chejan_slot] 매매 요청에 대한 주문 구분 : 잔고")
            print("[info] 잔고")
            account_num = self.dynamicCall("GetChejanData(int)", self.realType.REALTYPE['잔고']['계좌번호'])
            sCode = self.dynamicCall("GetChejanData(int)", self.realType.REALTYPE['잔고']['종목코드'])[1:]
            stock_name = self.dynamicCall("GetChejanData(int)", self.realType.REALTYPE['잔고']['종목명'])
            stock_name = stock_name.strip()
            current_price = self.dynamicCall("GetChejanData(int)", self.realType.REALTYPE['잔고']['현재가'])
            current_price = abs(int(current_price))
            stock_quan = self.dynamicCall("GetChejanData(int)", self.realType.REALTYPE['잔고']['보유수량'])
            stock_quan = int(stock_quan)
            like_quan = self.dynamicCall("GetChejanData(int)", self.realType.REALTYPE['잔고']['주문가능수량'])
            like_quan = int(like_quan)
            buy_price = self.dynamicCall("GetChejanData(int)", self.realType.REALTYPE['잔고']['매입단가'])
            buy_price = abs(int(buy_price))
            total_buy_price = self.dynamicCall("GetChejanData(int)", self.realType.REALTYPE['잔고']['총매입가'])
            total_buy_price = int(total_buy_price)
            meme_gubun = self.dynamicCall("GetChejanData(int)", self.realType.REALTYPE['잔고']['매도매수구분'])
            meme_gubun = self.realType.REALTYPE['매도수구분'][meme_gubun]
            first_sell_price = self.dynamicCall("GetChejanData(int)", self.realType.REALTYPE['잔고']['(최우선)매도호가'])
            first_sell_price = abs(int(first_sell_price))
            first_buy_price = self.dynamicCall("GetChejanData(int)", self.realType.REALTYPE['잔고']['(최우선)매수호가'])
            first_buy_price = abs(int(first_buy_price))

            if sCode not in self.jango_dict.keys():
                self.jango_dict.update({sCode: {}})

            # [반드시 해야 할 일]
            #
            # 01.수동 매매 성공시 self.portfolio_stock_dict에 update해야 함
            #   1. 이후 몇가지 정보를 업데이트 해야 함

            # 02. code.name.xlsx 파일을 생성해야 함
            #   1. datareader를 이용하여 엑셀 생성



            if sCode not in self.portfolio_stock_dict:  # 체결된 종목이 나의 관심 종목에 저장되어 있지 않다면
                self.portfolio_stock_dict.update({sCode: {}})
                print('[DEBUG] (수동으로 매매) ========================================================================')
                print('[DEBUG] self.portfolio_stock_dict : %s' % self.portfolio_stock_dict)
                self.logger.info("[DEBUG] (수동으로 매매) ========================================================================")
                self.logger.info("[DEBUG] self.portfolio_stock_dict : %s" % self.portfolio_stock_dict)

            self.jango_dict[sCode].update({"현재가": current_price})
            self.jango_dict[sCode].update({"종목코드": sCode})
            self.jango_dict[sCode].update({"종목명": stock_name})
            self.jango_dict[sCode].update({"보유수량": stock_quan})
            self.jango_dict[sCode].update({"주문가능수량": like_quan})
            self.jango_dict[sCode].update({"매입단가": buy_price})
            self.jango_dict[sCode].update({"총매입가": total_buy_price})
            self.jango_dict[sCode].update({"매도매수구분": meme_gubun})
            self.jango_dict[sCode].update({"(최우선)매도호가": first_sell_price})
            self.jango_dict[sCode].update({"(최우선)매수호가": first_buy_price})

            self.buyHistory.append(str(sCode))
            self.xlsxAdd(str(sCode), 'buyHistory.xlsx')

            if stock_quan == 0:
                del self.jango_dict[sCode]
                # 엑셀 파일 삭제
                # self.file_delete(sCode, stock_name)
                # 해당 종목에 대해 실시간 데이터 수신 대기 상태 해제

                #if sCode in self.portfolio_stock_dict.keys():
                #    self.dynamicCall("SetRealRemove(QString, QString)", self.portfolio_stock_dict[sCode]['스크린번호'], sCode)

    # 종목코드 가져오는 함수
    def get_code_list_by_market(self, market_code):
        code_list = self.dynamicCall("GetCodeListByMarket(QString)", market_code)
        code_list = code_list.split(";")[:-1]

        return code_list

    def calculator_fnc(self):
        # print("calculator_fnc")
        code_list = self.get_code_list_by_market("10")

        for idx, code in enumerate(code_list):
            self.dynamicCall("DisconnectRealData(QString)", self.screen_calculation_stock) # self.screen_calculation_stock : 스크린번호
            # DisconnectRealData : 해당하는 스크린 번호에 대한 실시간 데이터 받기를 해제한다.

            self.day_kiwoom_db(code=code)

    def day_kiwoom_db(self, code=None, date=None, sPrevNext="0"):
        print("day_kiwoom_db")
        QTest.qWait(3600) # 정보 요청을 3.6초 딜레이를 준다

        self.dynamicCall("SetInputValue(QString, QString", "종목코드", code)
        self.dynamicCall("SetInputValue(QString, QString", "수정주가구분", "1")

        if date != None:
            self.dynamicCall("SetInpuValue(QString, QString)", "기준일자", date)

        self.dynamicCall("CommRqData(QString, QString, int, QString)", "주식일봉차트조회", "opt10081", sPrevNext, self.screen_calculation_stock)
        self.calculator_event_loop.exec_()

    def read_code02(self):
        # 디렉토리 내의 모든 파일에 대해 루프를 돌면서 파일명 추출
        for filename in os.listdir("xlsx"):
            if os.path.isfile(os.path.join("xlsx", filename)):
                filecode = filename.split(".")[0]
                filenm = filename.split(".")[1]
                df = pd.read_excel('xlsx/' + filecode + '.' + filenm + '.xlsx', index_col=0, engine='openpyxl')
                self.portfolio_stock_dict.update({filecode:{}})
                self.portfolio_stock_dict[filecode].update({'Date':df['Date'].to_list(), 'High':df['High'].to_list(), 'Low':df['Low'].to_list(), 'Close':df['Close'].to_list()})
                # self.portfolio_stock_dict[filecode].update({'Ratio':df['Ratio'][0], '20MA':df['20MA'][0], 'BBUpper':df['BBUpper'][0]})

        # cci용 디렉토리 내의 모든 파일에 대해 루프를 돌면서 파일명 추출
        #for filename in os.listdir("xlsxConditions"):
        #    if os.path.isfile(os.path.join("xlsxConditions", filename)):
        #        ccifilecode = filename.split(".")[0]
        #        ccidf = pd.read_excel('xlsxConditions/' + ccifilecode + '.xlsx', engine='openpyxl')
        #        self.tmpStockDict.update({ccifilecode:{}})
        #        self.tmpStockDict[ccifilecode].update({'Date':ccidf['Date'].to_list(), 'High':ccidf['High'].to_list(), 'Low':ccidf['Low'].to_list(), 'Close':ccidf['Close'].to_list()})

        if os.path.isfile(os.path.join("historyXlsx", 'buyHistory.xlsx')):
            bh = pd.read_excel('historyXlsx/buyHistory.xlsx', dtype=str)
            if bh is not None:
                self.buyHistory = bh['buyHistory'].tolist()
                self.logger.info('[초기정보세팅] self.buyHistory : %s' % self.buyHistory)
                print('[초기정보세팅] self.buyHistory : %s' % self.buyHistory)
            else:
                # 파일을 읽는 데 문제가 있을 때의 처리
                print("Error: buyHistory - DataFrame is None, file may not be readable.")
        else:
            print("Error: buyHistory - file is not exist.")

        if os.path.isfile(os.path.join("historyXlsx", 'addBuyHistory.xlsx')):
            adb = pd.read_excel('historyXlsx/addBuyHistory.xlsx', dtype=str)
            if adb is not None:
                self.addBuyHistory = adb['addBuyHistory'].tolist()
                self.logger.info('[초기정보세팅] self.addBuyHistory : %s' % self.addBuyHistory)
                print('[초기정보세팅] self.addBuyHistory : %s' % self.addBuyHistory)
            else:
                # 파일을 읽는 데 문제가 있을 때의 처리
                print("Error: addBuyHistory - DataFrame is None, file may not be readable.")
        else:
            print("Error: addBuyHistory - file is not exist.")

        if os.path.isfile(os.path.join("historyXlsx", 'partialSalelist.xlsx')):
            ps = pd.read_excel('historyXlsx/partialSalelist.xlsx', dtype=str)
            if ps is not None:
                self.partialSalelist = ps['partialSalelist'].tolist()
                self.logger.info('[초기정보세팅] self.partialSalelist : %s' % self.partialSalelist)
                print('[초기정보세팅] self.partialSalelist : %s' % self.partialSalelist)
            else:
                # 파일을 읽는 데 문제가 있을 때의 처리
                print("Error: partialSalelist - DataFrame is None, file may not be readable.")
        else:
            print("Error: partialSalelist - file is not exist.")

    def screen_number_setting(self):
        # 각 dict에 겹치는 종목을 필터링한다.
        screen_overwrite = []

        #계좌평가잔고내역에 있는 종목들
        for code in self.account_stock_dict.keys():
            if code not in screen_overwrite:
                screen_overwrite.append(code)

        #미체결에 있는 종목들
        for order_number in self.not_account_stock_dick.keys():
            code = self.not_account_stock_dick[order_number]['종목코드']
            if code not in screen_overwrite:
                screen_overwrite.append(code)

        #포트폴리오에 담겨있는 종목들
        for code in self.portfolio_stock_dict.keys():
            if code not in screen_overwrite:
                screen_overwrite.append(code)

        #스크린 번호 할당
        cnt = 0
        for code in screen_overwrite:
            temp_screen = int(self.screen_real_stock)
            meme_screen = int(self.screen_meme_stock)

            if (cnt % 50) == 0:
                temp_screen += 1
                self.screen_real_stock = str(temp_screen)

            if (cnt % 50) == 0:
                meme_screen += 1
                self.screen_meme_stock = str(meme_screen)

            if code in self.portfolio_stock_dict.keys():
                self.portfolio_stock_dict[code].update({"스크린번호": str(self.screen_real_stock)})
                self.portfolio_stock_dict[code].update({"주문용스크린번호": str(self.screen_meme_stock)})

            elif code not in self.portfolio_stock_dict.keys():
                print('========================================================================================================================')
                self.portfolio_stock_dict.update({code: {"스크린번호":str(self.screen_real_stock), "주문용스크린번호": str(self.screen_meme_stock)}})

            cnt += 1

    # 송수신 메시지 set
    def msg_slot(self, sScrNo, sRQName, sTrCode, msg):
            self.logger.info("msg_slot() => 메시지 요청에 대한 응답 / 스크린 : %s, 요청이름 : %s, tr코드 : %s --- %s" % (sScrNo, sRQName, sTrCode, msg))
            # print("msg_slot() => 메시지 요청에 대한 응답 / 스크린 : %s, 요청이름 : %s, tr코드 : %s --- %s" % (sScrNo, sRQName, sTrCode, msg))

    # 파일 삭제
    def file_delete(self, sCode, stock_name):
        if os.path.isfile("xlsx/"+sCode+"."+stock_name+".xlsx"):
            os.remove("xlsx/"+sCode+"."+stock_name+".xlsx")

    # 신규/추가 매수, 부분매도를 엑셀로 저장하는 함수
    def historySaveExcel(self, data_list, file_name):
        # 디렉토리 생성
        directory = 'historyXlsx'
        if not os.path.exists(directory):
            os.makedirs(directory)

        # 엑셀 파일 경로
        file_path = os.path.join(directory, file_name + '.xlsx')

        # 새로운 Workbook 생성
        workbook = Workbook()

        # 첫 번째 시트 선택
        sheet = workbook.active

        # 컬럼명 설정
        sheet['A1'] = 'Index'
        sheet['B1'] = file_name

        # 데이터 추가
        for index, data in enumerate(data_list):
            print('index:%s' % index)
            sheet[f'A{index + 2}'] = index
            sheet[f'B{index + 2}'] = data

        # 변경 사항을 저장
        workbook.save(file_path)

    def clearAndSaveExcel(self, data_list, file_name):
        # 디렉토리 생성
        directory = 'historyXlsx'
        if not os.path.exists(directory):
            os.makedirs(directory)

        # 엑셀 파일 경로
        file_path = os.path.join(directory, file_name + '.xlsx')

        # 기존 엑셀 파일이 있는 경우 내용 삭제 후 열기
        if os.path.exists(file_path):
            wb = load_workbook(file_path)
            ws = wb.active
            # 모든 셀 내용 삭제
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.value = None
            # 변경 사항 저장
            wb.save(file_path)
        else:
            # 새로운 Workbook 생성
            workbook = Workbook()
            # 첫 번째 시트 선택
            sheet = workbook.active
            # 컬럼명 설정
            sheet['A1'] = 'Index'
            sheet['B1'] = file_name
            workbook.save(file_path)

        # 새로운 데이터 추가
        workbook = load_workbook(file_path)
        sheet = workbook.active
        for index, data in enumerate(data_list):
            sheet[f'A{index + 2}'] = index
            sheet[f'B{index + 2}'] = data
        # 변경 사항 저장
        workbook.save(file_path)


    def xlsxDelete(self, sCode, fName):
        # 엑셀 파일 불러오기
        workbook = load_workbook(fName)

        # 첫 번째 시트 선택
        sheet = workbook.active

        # 'buyHistory' 컬럼 내에서 sCode가 있는 행 찾기
        for idx, row_data in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            print('비교 - %s : %s' % (row_data[1], sCode))
            if row_data[1] == sCode:  # 'buyHistory' 컬럼의 인덱스가 0이라고 가정
                # 해당 행 삭제
                print(row_data[1])
                sheet.delete_rows(idx)

        # 변경된 내용을 저장
        workbook.save(fName)

    def xlsxAdd(self, sCode, fName):
        print('sCode : %s' % sCode)

        # 엑셀 파일 경로 및 파일명
        file_path = self.filePath + fName

        # 엑셀 파일 불러오기
        workbook = load_workbook(file_path)

        # 첫 번째 시트 선택
        sheet = workbook.active

        # 'buyHistory' 컬럼의 마지막 행 인덱스 찾기
        last_row_index = sheet.max_row + 1

        # 새로운 sCode 추가
        sheet.cell(row=last_row_index, column=2, value=sCode)

        for idx, row_data in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            print('비교 - %s : %s' % (row_data[1], sCode))

        # 변경된 내용을 저장
        workbook.save(fName)

    def historyListLogging(self, code_nm, sCode, flag):
        self.logger.info("[%s(%s) - %s 발생] ☆☆☆☆☆☆☆☆☆☆☆☆☆☆☆☆☆☆☆☆☆☆☆☆☆☆☆☆☆☆☆☆" % (code_nm, sCode, flag))
        self.logger.info("매수 이력 딕셔너리 : %s" % self.buyHistory)
        self.logger.info("추가 매수 이력 딕셔너리 : %s" % self.addBuyHistory)
        self.logger.info("부분매도 이력 딕셔너리 : %s" % self.partialSalelist)

        print("[%s(%s) - %s 발생] ☆☆☆☆☆☆☆☆☆☆☆☆☆☆☆☆☆☆☆☆☆☆☆☆☆☆☆☆☆☆☆☆" % (code_nm, sCode, flag))
        print("매수 이력 딕셔너리 : %s" % self.buyHistory)
        print("추가 매수 이력 딕셔너리 : %s" % self.addBuyHistory)
        print("부분매도 이력 딕셔너리 : %s" % self.partialSalelist)

    def getMaPrice(self, sCode, ndays=20):
        print(self.portfolio_stock_dict.keys())
        data = self.portfolio_stock_dict[sCode]
        # pd.set_option('display.max_rows', None)
        # pd.set_option('display.max_columns', None)
        print('sCode: %s' % sCode)
        data = pd.DataFrame(data)
        # print(data)
        data = data.tail(50)

        #20일 평균 가격
        ma20_price = data['Close'].rolling(window=20).mean()
        data['MA20'] = ma20_price

        rtnVal = data['MA20'].iloc[-1]
        data.drop(data.index[-1])
        #print('▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒')
        #print('High : %s | Low : %s | Close : %s' % (highVal, lowVal, closeVal))
        #print(data)
        #print('[DEBUG][평균값구하기] 종목코드 : %s | 20일평균 : %s | 현재가 : %s | 날짜 : %s' % (sCode, data['MA20'].iloc[-1], closeVal, datetime.today().date()))
        #if len(ma20_price) > 0 and data["Date"].iloc[-1] == datetime.today().date():
        #    #rtnVal = True
        #    return data['MA20'].iloc[-1]
        #else:
        #    # DataFrame의 마지막 행 제거
        #    data.drop(data.index[-1])
        #    return False
        return rtnVal

    '''
    시그널 발새이후
    1. 시그널 일자보다 낮지 않아야 한다.
    2. 50일 선을 지지해야 한다.
    getCCI를 통해 Dataframe을 구한다.
    마지막 시그널 일자 이후 종가를 구한다.
    종가가 50일 선에
    '''

    def getCCI(self, highVal, lowVal, closeVal, sCode, ndays=50):
        rtnVal = False

        data = self.portfolio_stock_dict[sCode]
        data = pd.DataFrame(data)

        # 신규 삽입할 데이터
        new_data = {'Date': datetime.today().date(), 'High': highVal, 'Low': lowVal, 'Close': closeVal}

        # DataFrame에 새로운 행 삽입
        data.loc[len(data)] = new_data

        tp = (data['High'] + data['Low'] + data['Close']) / 3
        tp_rolling_mean = tp.rolling(window=50).mean()
        mean_deviation = tp.rolling(window=50).std()

        cciData = (tp - tp_rolling_mean) / (0.015 * mean_deviation)

        data['CCI'] = cciData

        # Crossup 신호 검색
        crossup_signals = data[(data['CCI'].shift(1) < 0) & (data['CCI'] > 0)]

        if len(crossup_signals) > 0 and crossup_signals["Date"].iloc[-1] == datetime.today().date():
            print('crossup_signals : %s' % crossup_signals)
            rtnVal = True
        else:
            # DataFrame의 마지막 행 제거
            pass

        data.drop(data.index[-1])

        return rtnVal

    # 보유 종목 최대 4개 제한 + 예비비 = 총 5개
    # 종목별 예산 할당
    # 종목의 단계별 매수 가능액 할당
    #